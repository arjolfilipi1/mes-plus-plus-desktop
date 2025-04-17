from peewee import *
import datetime
import openpyxl

db = SqliteDatabase('1.db')

class BaseModel(Model):
    class Meta:
        database = db
level =(
('test','test'),
('AV','AV'),
('cutting','cutting'),
('assy','assy'),

)
class UserLevel(BaseModel):
    name     = CharField(primary_key=True)
    Parent   = ForeignKeyField('self', to_field='name',null=True, backref='children',on_delete='CASCADE')
    rank     = IntegerField(default = 0)

class Departament(BaseModel):
    name     = CharField(primary_key=True)
    Parent   = ForeignKeyField('self', to_field='name',null=True, backref='children',on_delete='CASCADE')

class Category(BaseModel):
    name     = CharField(primary_key=True)
    Parent   = ForeignKeyField('self', to_field='name',null=True, backref='children',on_delete='CASCADE')
    terminal = BooleanField(default = False)
    description = CharField(  null = True, max_length = 550 )
    # Update the category of all the children to be the same
    def make_children_terminal(self,value):
        children = self.children
        for child in children:
            child.terminal = True
            child.save()
    def save(self, *args, **kwargs):
        self.make_children_terminal(self.terminal)
        return super(Category, self).save(*args, **kwargs)
       
class UserPermision(BaseModel):
    UserLevel      = ForeignKeyField(UserLevel, to_field="name",on_delete='CASCADE')
    permision_name = CharField(max_length=250)
    permision      = BooleanField(default= False)
class User(BaseModel):
    username    = CharField(unique=True)
    operator_id = CharField()
    password    = CharField()
    op_level    = ForeignKeyField(UserLevel, null= True,to_field="name",on_delete='SET NULL')
    
class StockLocationType(BaseModel):
    name        = CharField(unique=True,max_length=100,primary_key=True)
    description = CharField(unique=True,max_length=250)
# class StockLocation(BaseModel):
    
class StockLocation(BaseModel):
    name        = CharField(primary_key = True)
    Owner       = ForeignKeyField(User, null= True,to_field="id",on_delete='SET NULL')
    SL_type     = ForeignKeyField(StockLocationType, to_field="name",on_delete='CASCADE')
    Departament = ForeignKeyField(Departament,null=True, to_field="name",on_delete='SET NULL')
    Structural  = BooleanField(default =  False)
    
class Part_nr(BaseModel):
    Part_nr_txt  = CharField(primary_key = True)
    Drawing_nr   = CharField(null = True)
    Draving_date = DateField(null = True)
    Virtual      = BooleanField(default = False)
    Selable      = BooleanField(default = False)
    Purchaseble  = BooleanField(default = False)
    Active       = BooleanField(default = False)
    Tracked      = BooleanField(default = False)
    Change_index = CharField(null = True)
    Note         = CharField(null = True)
    Safety_stock = FloatField(default = 0)
    Def_location = ForeignKeyField( StockLocation,null=True, to_field="name", on_delete='SET NULL', backref='children')
    Category  = ForeignKeyField( Category,null=True, to_field="name", on_delete='SET NULL')
    # 1	444000016176	6029027087	17.10.2017	C
    
class StockItem(BaseModel):
    Parent   = ForeignKeyField('self', null=True, backref='children',on_delete='SET NULL')
    Part     = ForeignKeyField(Part_nr, to_field="Part_nr_txt",on_delete='CASCADE', backref='children')
    Location = ForeignKeyField(StockLocation, to_field="id",on_delete='CASCADE', backref='children')
    
    
class CuttingList(BaseModel):
    #name of single wire, to be used on label if it is needed to be specified
    HProduct_part_nr = CharField(null = True)
    #Option or module
    Opt              = CharField(null = True)
    # Unique for each wire for this part number
    Position         = CharField(null = True)
    # text for wire printing
    Print            = CharField(null = True)
    # wire material
    Material         = CharField(null = True)
    # cross sectional area in sqmm
    Sq               = CharField(null = True)
    # color code, to be used in vizualization
    Color            = CharField(null = True)
    # first endpoint
    From             = CharField(null = True)
    # position on first connector etc
    Pin_L            = CharField(null = True)
    # kontakt to be used
    Kontakt_L        = CharField(null = True)
    # seal to be used
    Seal_L           = CharField(null = True)
    # strip
    Strip_L          = FloatField(null = True)
    # used for instertion hint
    Adress_L         = CharField(null = True)
    # tool to be used, usualy applikator
    Tool_L           = CharField(null = True)
    To               = CharField(null = True)
    Pin_R            = CharField(null = True)
    Kontakt_R        = CharField(null = True)
    Seal_R           = CharField(null = True)
    Strip_R          = FloatField(null = True)
    Adress_R         = CharField(null = True)
    Tool_R           = CharField(null = True)
    # wire cut length
    Length           = FloatField(null = True)
    # dedicated machine, can be left empty for auto allocation
    Machine          = CharField(null = True)
    # work time
    Time             = FloatField(null = True)
    # wire part number
    Wire             = CharField(null = True)
    # packing bundle, also used for ticket printing
    Bundle           = IntegerField(null = True)
    # parent part number
    part             = ForeignKeyField(Part_nr, to_field="Part_nr_txt",on_delete='CASCADE', backref='children')
    
# class BOM(BaseModel):
    
# db.connect()
# suser = Category(name='terminal',Parent= None,terminal = False )
# suser.save()

# print(suser.name)
# db.create_tables([Category])
def Excel_empty(model,location =''):
    field_names = list(model._meta.columns.keys())
    tname = model.__name__
    filepath = location+tname+".xlsx"
    wb = openpyxl.Workbook()
    i =0
    ws = wb.active
    while i < len(field_names):
        ws.cell(row=1, column=i+1).value = field_names[i]
        i +=1
    try:
        wb.save(filepath)
        return True, True
    except Exception as e:
        return False , e 

# Excel_empty(UserLevel)

#automaticly add data from filled excel to the database
def ReadExcel(model,file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    success = False
    i = 0
    headers=[]
    skip = []
    to_add=[]
    for row in ws:
        data={}
        if i ==0:
            for c in row:
                headers.append(c.value)
                if c.value in ['id']:
                    skip.append(c.value)
        else:
            j = 0
            for c in row:
                if c.value not in skip:
                    data[headers[j]] = c.value
                j+=1
            to_add.append(data)
        i+=1

    try:
        with db.atomic():
            r = model.insert_many(to_add).execute()
        return True
    except Exception as e:
        return success,e
# ReadExcel(UserLevel,'UserLevel.xlsx')
# print(datetime.datetime.now().date())